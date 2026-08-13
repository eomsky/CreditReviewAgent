"""Conversational endpoint backed by the multimodal Colab vLLM server."""

import base64
import binascii
import io
import mimetypes
import re
import uuid
from pathlib import Path
from typing import Any

import httpx
from fastapi import APIRouter, HTTPException
from pydantic import BaseModel, Field
from pypdf import PdfReader

from app.clients.colab_llm import ColabLLMClient
from app.core.config import settings


router = APIRouter()


class ChatMessage(BaseModel):
    role: str = Field(pattern="^(system|user|assistant)$")
    content: str = Field(min_length=1, max_length=50_000)


class ChatAttachment(BaseModel):
    filename: str = Field(min_length=1, max_length=255)
    mime_type: str = Field(min_length=1, max_length=100)
    data_base64: str = Field(min_length=1)


class ChatRequest(BaseModel):
    messages: list[ChatMessage] = Field(min_length=1, max_length=100)
    attachments: list[ChatAttachment] = Field(default_factory=list, max_length=10)


class ChatResponse(BaseModel):
    message: ChatMessage


@router.post("/completions", response_model=ChatResponse)
async def create_chat_completion(request: ChatRequest) -> ChatResponse:
    messages: list[dict[str, Any]] = [message.model_dump() for message in request.messages]
    if not any(message["role"] == "system" for message in messages):
        messages.insert(
            0,
            {
                "role": "system",
                "content": (
                    "당신은 금융기관의 신중한 여신심사 보조 에이전트입니다. "
                    "근거가 부족한 내용은 추측하지 말고 추가 자료를 요청하세요."
                ),
            },
        )

    if request.attachments:
        await _attach_files_to_last_user_message(messages, request.attachments)

    try:
        content = await ColabLLMClient().complete(messages)
    except httpx.HTTPStatusError as exc:
        raise HTTPException(status_code=502, detail="Colab LLM이 요청을 거절했습니다.") from exc
    except (httpx.RequestError, KeyError, TypeError) as exc:
        raise HTTPException(status_code=503, detail="Colab LLM 서버에 연결할 수 없습니다.") from exc

    return ChatResponse(message=ChatMessage(role="assistant", content=content))


async def _attach_files_to_last_user_message(
    messages: list[dict[str, Any]], attachments: list[ChatAttachment]
) -> None:
    image_count = sum(item.mime_type.startswith("image/") for item in attachments)
    if image_count > settings.MAX_IMAGES_PER_MESSAGE:
        raise HTTPException(
            status_code=422,
            detail=(
                f"한 번에 이미지는 최대 {settings.MAX_IMAGES_PER_MESSAGE}개까지 분석할 수 있습니다. "
                "나머지 이미지는 다음 메시지에 첨부해 주세요."
            ),
        )

    user_message = next(
        (message for message in reversed(messages) if message["role"] == "user"), None
    )
    if user_message is None:
        raise HTTPException(status_code=422, detail="첨부파일에는 사용자 메시지가 필요합니다.")

    content: list[dict[str, Any]] = [{"type": "text", "text": user_message["content"]}]
    extracted_documents: list[str] = []

    for attachment in attachments:
        raw = _decode_attachment(attachment)
        stored_path = _store_attachment(attachment.filename, raw)

        if attachment.mime_type.startswith("image/"):
            content.append(
                {
                    "type": "image_url",
                    "image_url": {
                        "url": f"data:{attachment.mime_type};base64,{base64.b64encode(raw).decode('ascii')}"
                    },
                }
            )
        else:
            text = _extract_document_text(stored_path, attachment.mime_type, raw)
            extracted_documents.append(
                f"\n--- 첨부 문서: {attachment.filename} ---\n{text or '[추출 가능한 텍스트 없음]'}"
            )

    if extracted_documents:
        content[0]["text"] += (
            "\n\n다음은 첨부 문서에서 추출한 내용입니다. 문서에 없는 내용은 추측하지 마세요."
            + "".join(extracted_documents)
        )[: settings.MAX_DOCUMENT_TEXT_CHARS]
    user_message["content"] = content


def _decode_attachment(attachment: ChatAttachment) -> bytes:
    try:
        raw = base64.b64decode(attachment.data_base64, validate=True)
    except (binascii.Error, ValueError) as exc:
        raise HTTPException(status_code=422, detail=f"잘못된 첨부파일: {attachment.filename}") from exc
    if len(raw) > settings.MAX_UPLOAD_SIZE_MB * 1024 * 1024:
        raise HTTPException(
            status_code=413,
            detail=f"{attachment.filename}: 최대 {settings.MAX_UPLOAD_SIZE_MB}MB까지 업로드할 수 있습니다.",
        )
    return raw


def _store_attachment(filename: str, raw: bytes) -> Path:
    safe_name = re.sub(r"[^0-9A-Za-z가-힣._-]", "_", Path(filename).name)
    target = settings.UPLOAD_DIR / f"{uuid.uuid4().hex}_{safe_name}"
    target.write_bytes(raw)
    return target


def _extract_document_text(path: Path, mime_type: str, raw: bytes) -> str:
    suffix = path.suffix.lower()
    if mime_type == "application/pdf" or suffix == ".pdf":
        reader = PdfReader(io.BytesIO(raw))
        return "\n".join(page.extract_text() or "" for page in reader.pages)
    if suffix == ".docx":
        from docx import Document

        document = Document(io.BytesIO(raw))
        return "\n".join(paragraph.text for paragraph in document.paragraphs)
    if suffix in {".xlsx", ".xlsm"}:
        from openpyxl import load_workbook

        workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        lines: list[str] = []
        for sheet in workbook.worksheets:
            lines.append(f"[시트: {sheet.title}]")
            for row in sheet.iter_rows(values_only=True):
                lines.append("\t".join("" if value is None else str(value) for value in row))
        return "\n".join(lines)
    if suffix in {".txt", ".csv", ".md", ".json"} or mime_type.startswith("text/"):
        return raw.decode("utf-8", errors="replace")
    guessed, _ = mimetypes.guess_type(path.name)
    raise HTTPException(
        status_code=415,
        detail=f"현재 텍스트 추출을 지원하지 않는 문서 형식입니다: {guessed or suffix}",
    )
